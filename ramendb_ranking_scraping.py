import time
import pandas as pd
import os
import csv
import random
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException, TimeoutException, WebDriverException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font

# ファイルパス設定
EXCEL_PATH = "C:/Users/kanie/OneDrive/ramendb_ranking_scraping.xlsx"
CSV_PATH = "C:/Users/kanie/OneDrive/ranking_temp.csv"

# ログイン情報
EMAIL = "kanieksuke@yahoo.co.jp"
PASSWORD = "chimpo"

# Chromeオプション設定（アクセス制限回避）
chrome_options = Options()
# chrome_options.add_argument("--headless")
chrome_options.add_argument("--disable-blink-features=AutomationControlled")  # ボット検出回避
# chrome_options.add_argument("--incognito")
chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")  # User-Agent偽装

# ChromeDriverの起動
service = Service()  # ChromeDriverのデフォルトパスを使用
driver = webdriver.Chrome(service=service, options=chrome_options)
driver.set_page_load_timeout(10)

def sleep():
    time.sleep(random.uniform(3, 5))
    # time.sleep(1)

# ラーメンデータベースのトップページにアクセス
url = 'https://ramendb.supleks.jp/'
print(f'Accessing: {url}')
driver.get(url)
sleep()

# ログインページへ遷移
try:
    login_link = driver.find_element(By.LINK_TEXT, "ログイン")
    login_link.click()
    print("Navigated to login page.")
    sleep()  # ページ遷移後の待機
except NoSuchElementException:
    print("Login link not found. Exiting.")
    driver.quit()
    exit()

# メールアドレスとパスワードを入力
try:
    email_input = driver.find_element(By.NAME, "mail")
    email_input.send_keys(EMAIL)
    
    password_input = driver.find_element(By.NAME, "password")
    password_input.send_keys(PASSWORD)
    
    print("Entered login credentials.")

    # ログインボタンをクリック
    login_button = driver.find_element(By.CLASS_NAME, "formbtn")
    login_button.click()
    
    print("Clicked login button.")
    sleep()  # ログイン処理の待機
except NoSuchElementException:
    print("Login form not found. Exiting.")
    driver.quit()
    exit()

print("Login successful")

# 東京版ラーメンデータベースへのリンクをクリック
try:
    ramen_link = driver.find_element(By.XPATH, "//a[@name='ramendb' and contains(@href, 'tokyo-ramendb.supleks.jp')]")
    driver.execute_script("arguments[0].scrollIntoView();", ramen_link)
    driver.execute_script("arguments[0].click();", ramen_link)
    # print("Clicked Tokyo_ramen_db link. Waiting for next page.")
    # WebDriverWait(driver, 10).until(
    #     EC.presence_of_element_located((By.LINK_TEXT, "ランキング"))
    # )
    print("Navigated to Tokyo Ramen Database page.")
    sleep()  # ページ遷移後の待機
except NoSuchElementException:
    print("Link not found. Exiting.")

# ランキングへのリンクをクリック
try:
    rank_link = driver.find_element(By.LINK_TEXT, "ランキング")
    driver.execute_script("arguments[0].scrollIntoView();", rank_link)
    driver.execute_script("arguments[0].click();", rank_link)
    # print("Clicked ranking link. Waiting for next page.")
    # WebDriverWait(driver, 10).until(
    #     EC.presence_of_element_located((By.LINK_TEXT, "お店注目度"))
    # )
    print("Navigated to Ranking page.")
    sleep()  # ページ遷移後の待機
except NoSuchElementException:
    print("Link not found. Exiting.")

# お店注目度ランキングへのリンクをクリック
try:
    hot_shop_link = driver.find_element(By.XPATH, "//a[contains(@href, '/rank/hot-shop')]//span[contains(text(), 'お店注目度')]")
    driver.execute_script("arguments[0].scrollIntoView();", hot_shop_link)
    driver.execute_script("arguments[0].click();", hot_shop_link)
    print("Navigated to Hot shop page.")
    sleep()  # ページ遷移後の待機
except NoSuchElementException:
    print("Link not found. Exiting.")

# 結果を保存するリスト
results = []

# nameクラスのリンクをすべて取得
name_elements = driver.find_elements(By.CLASS_NAME, "name")

print(f"Found {len(name_elements)} entries.")

# 100件分ループ（要素は都度取得し直す）
for i in range(100):
    try:
        # nameクラスのリンクはページ遷移で消えるため、その都度取得し直す
        name_elements = driver.find_elements(By.CLASS_NAME, "name")
        elem = name_elements[i]

        # リンクを新たに取得しクリック
        driver.execute_script("arguments[0].scrollIntoView();", elem)
        elem.click()
        sleep()

        # 遷移先URL
        current_url = driver.current_url

        # 店名の取得
        try:
            shop_name = driver.find_element(By.XPATH, "//th[text()='店名']/following-sibling::td").text
        except NoSuchElementException:
            shop_name = "店名取得失敗"

        # 現在の日付を取得
        today = datetime.today().date()

        # 初期化
        open_date_raw = ""
        open_date_obj = None
        open_date_str = ""

        # オープン日取得・変換
        try:
            td_element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//th[text()='オープン日']/following-sibling::td"))
            )
            open_date_raw = td_element.text.strip()

            if open_date_raw and "年" in open_date_raw and "月" in open_date_raw and "日" in open_date_raw:
                open_date_obj = datetime.strptime(open_date_raw, "%Y年%m月%d日").date()
                open_date_str = open_date_obj.strftime("%Y/%m/%d")
            else:
                open_date_str = open_date_raw
        except:
            open_date_raw = "オープン日取得失敗"
            open_date_str = open_date_raw

        # 条件1: class="btn-bookmark off"が存在するか
        bookmark_off_exists = False
        try:
            driver.find_element(By.CLASS_NAME, "btn-bookmark.off")
            bookmark_off_exists = True
        except NoSuchElementException:
            pass

        # 条件2: オープン日が90日以内か
        open_within_90_days = False
        if open_date_obj:
            days_diff = (today - open_date_obj).days
            open_within_90_days = 0 <= days_diff <= 90  # ※未来日も除外する

        # 両方の条件を満たす場合のみ追加
        if bookmark_off_exists and open_within_90_days:
            results.append({
                "クリック回数": i + 1,
                "URL": driver.current_url,
                "店名": shop_name,
                "オープン日": open_date_str
            })
            print(f"✅ 追加: {shop_name} ({open_date_str})")
        else:
            print(f"⏩ スキップ: {shop_name} - 条件不一致")

        # ランキングページに戻る
        driver.back()
        sleep()

    except Exception as e:
        print(f"{i+1}件目でエラー発生: {e}")
        break

# CSVに保存
df = pd.DataFrame(results)
df.to_csv(CSV_PATH, index=False, encoding="utf-8-sig")
print(f"{len(results)}件のデータをCSVに保存しました: {CSV_PATH}")

# input("ブラウザを手動で操作した後、Enter を押してください...") 

# 日付文字列の作成（1行目に記録する）
today_str = datetime.today().strftime("%Y年%m月%d日")

# Excelファイルの読み込み
wb = load_workbook(EXCEL_PATH)
ws = wb.active

# --- ① 出力先列（新しい日付列）を決める ---
date_col = None
for col in range(2, ws.max_column + 2):  # B列(2)から順に
    if ws.cell(row=1, column=col).value is None:
        ws.cell(row=1, column=col).value = today_str
        date_col = col
        break

if date_col is None:
    raise Exception("空いている日付列が見つかりません")

# --- ② CSVデータの読み込み ---
with open(CSV_PATH, newline='', encoding='utf-8-sig') as f:
    reader = csv.DictReader(f)
    csv_data = list(reader)

# --- ③ CSVデータをExcelに反映 ---
updated_rows = []  # 並べ替え対象の行番号を記録

for row in csv_data:
    shop_name = row["店名"]
    click_count = int(row["クリック回数"])

    # A列（店名列）から該当店の行を探す
    matched_row = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == shop_name:
            matched_row = r
            break

    if matched_row is None:
        # 空行に追加
        for r in range(2, ws.max_row + 2):
            if ws.cell(row=r, column=1).value is None:
                ws.cell(row=r, column=1).value = shop_name
                matched_row = r
                break

    # 指定列にクリック回数を入力
    target_cell = ws.cell(row=matched_row, column=date_col)
    target_cell.value = click_count

    # 左隣のクリック回数と比較
    prev_cell = ws.cell(row=matched_row, column=date_col - 1)
    try:
        prev_val = int(prev_cell.value)
        if click_count > prev_val:
            target_cell.font = Font(color="0000FF")  # 青
        elif click_count < prev_val:
            target_cell.font = Font(color="FF0000")  # 赤
        # 同値は色なし
    except (TypeError, ValueError):
        target_cell.font = Font(color="FF0000")  # 前日が空欄など

    updated_rows.append((matched_row, click_count))

# --- ④ 並び替え処理（クリック回数昇順） ---
# 並び替えるデータだけ抽出・ソート
sorted_rows = sorted(updated_rows, key=lambda x: x[1])

# 並び替え対象外の行（今回更新されてない店）を後ろに
other_rows = [
    r for r in range(2, ws.max_row + 1)
    if r not in [x[0] for x in updated_rows]
]

# 並び替え後の行データを一時リストに記録
all_sorted_rows = sorted_rows + [(r, None) for r in other_rows]
new_rows = []

for row_idx, _ in all_sorted_rows:
    row_data = [ws.cell(row=row_idx, column=col).value for col in range(1, ws.max_column + 1)]
    new_rows.append(row_data)

# 旧データをクリア
for r in range(2, ws.max_row + 1):
    for c in range(1, ws.max_column + 1):
        ws.cell(row=r, column=c).value = None
        ws.cell(row=r, column=c).font = Font(color="000000")  # フォント色をリセット

# 新しい順序で書き戻し
for i, row_data in enumerate(new_rows, start=2):
    for j, val in enumerate(row_data, start=1):
        ws.cell(row=i, column=j).value = val

# --- ⑤ 保存＆CSV削除 ---
wb.save(EXCEL_PATH)
wb.close()
os.remove(CSV_PATH)

print(f"Excelファイルを更新し、CSVファイルを削除しました：{EXCEL_PATH}")

driver.quit()

